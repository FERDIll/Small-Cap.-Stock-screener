#!/usr/bin/env python3
"""
SEC small-cap screener scraper (gated, failure-tolerant, NO TIERS, NO "in-depth" layer)

Design decisions implemented per your latest spec:
- Visible, concise sheet structure: Intro + Overview, with everything else pushed into "Clutter".
- SIC code is treated as clutter (still retained for traceability / mapping).
- Price/quote mechanics are clutter (close price, price date, $ADV).
- No Tier A/B/C expansions; no deep parsing beyond what is required for Gate 2 + concise Overview.
- Gate 3 still fetches Yahoo LAST (single chart call), but only MarketCap + ADV30_Shares are “Overview”.
"""

from __future__ import annotations

import csv
import json
import re
import time
import zipfile
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from xml.etree import ElementTree as ET

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

SEC_SESSION = requests.Session()
YAHOO_SESSION = requests.Session()


# -----------------------------
# Configuration
# -----------------------------

REPO_ROOT = Path(__file__).resolve().parents[1]

DATA_DIR = REPO_ROOT / "data"
CONFIG_DIR = REPO_ROOT / "config"
CACHE_DIR = REPO_ROOT / ".cache" / "sec"

DATA_DIR.mkdir(parents=True, exist_ok=True)
CONFIG_DIR.mkdir(parents=True, exist_ok=True)
CACHE_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_XLSX = DATA_DIR / "defense_screening_prototype_v1.xlsx"

UNIVERSE_CANDIDATES = [
    DATA_DIR / "universe.csv",
    DATA_DIR / "Universe.xlsx",
    DATA_DIR / "universe.xlsx",
    OUTPUT_XLSX,  # fallback: tickers may be in the Universe sheet of the output workbook
]

UNIVERSE_SHEET = "Universe"
UNIVERSE_TICKER_HEADER = "ticker"  # case-insensitive

# Optional SIC config (not a hard gate)
SIC_CONFIG = CONFIG_DIR / "sic_aero_defense.json"
DEFAULT_SIC_ALLOWLIST = [3480, 3720, 3721, 3724, 3728, 3760, 3812]

SEC_COMPANY_TICKERS_URL = "https://www.sec.gov/files/company_tickers.json"
SEC_SUBMISSIONS_URL_TMPL = "https://data.sec.gov/submissions/CIK{cik10}.json"
SEC_COMPANYFACTS_URL_TMPL = "https://data.sec.gov/api/xbrl/companyfacts/CIK{cik10}.json"

SEC_ARCHIVES_PRIMARYDOC_TMPL = (
    "https://www.sec.gov/Archives/edgar/data/{cik_nolead}/{acc_nodash}/{primary_doc}"
)

# IMPORTANT: SEC requests a real UA with contact info
USER_AGENT = "Ferdinand Niggemeier Small-Cap Stock Screener (Niggemeier.Ferdinand@gmail.com)"

SEC_SECONDS_BETWEEN_REQUESTS = 0.20
YAHOO_SECONDS_BETWEEN_REQUESTS = 0.00
HTTP_TIMEOUT = 30

# Gate 1 windows (tighter)
G1_MAX_AGE_10K_DAYS = 15 * 30  # ~15 months
G1_MAX_AGE_10Q_DAYS = 6 * 30   # ~6 months
G1_MAX_AGE_20F_DAYS = 18 * 30  # ~18 months

ENABLE_GATE_1 = True
ENABLE_GATE_2 = True
ENABLE_GATE_3 = True

# Gate 2 (tight, EDGAR-only)
G2_MIN_TTM_REVENUE_USD = 25_000_000
G2_MIN_TOTAL_ASSETS_USD = 50_000_000
G2_MIN_CASH_USD = 5_000_000

# Gate 3 (fetched last)
G3_MAX_MARKET_CAP_USD = 5_000_000_000
G3_MIN_ADV30_SHARES = 200_000
G3_MIN_ADV30_DOLLAR = 2_000_000


# -----------------------------
# Helpers: basic utils
# -----------------------------

def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


TICKER_RE = re.compile(r"^[A-Z0-9][A-Z0-9.\-]{0,9}$")  # allows BRK.B, RDS-A, etc.


def clean_ticker(raw: str) -> Optional[str]:
    t = (raw or "").strip().upper()
    if not t:
        return None
    if not TICKER_RE.match(t):
        return None
    return t


def ticker_variants(t: str) -> List[str]:
    t = (t or "").strip().upper()
    if not t:
        return []
    variants = [t, t.replace(".", "-"), t.replace("-", ".")]
    out: List[str] = []
    seen = set()
    for x in variants:
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def sleep_sec() -> None:
    time.sleep(SEC_SECONDS_BETWEEN_REQUESTS)

def sleep_yahoo() -> None:
    time.sleep(YAHOO_SECONDS_BETWEEN_REQUESTS)


def parse_date(s: str) -> Optional[date]:
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def na(v: Any) -> Any:
    if v is None:
        return "N/A"
    if isinstance(v, str) and not v.strip():
        return "N/A"
    return v


def safe_float(v: Any) -> Optional[float]:
    try:
        if v is None:
            return None
        return float(v)
    except Exception:
        return None


def normalize_header(s: str) -> str:
    return " ".join(str(s).strip().lower().split())


# -----------------------------
# SIC -> Sector mapping (simple, fast, edit as you wish)
# -----------------------------

DEFENSE_SIC = {3480, 3720, 3721, 3724, 3728, 3760, 3812}

def sic_to_sector(sic: Optional[int]) -> Optional[str]:
    if sic is None:
        return None
    # Your preferred mapping (tight, transparent)
    if sic in {3720, 3721, 3724, 3728, 3760}:
        return "Aerospace & Defense"
    if sic in {3480, 3812}:
        return "Defense Components"
    # Fallback coarse bucket by SIC major group
    first = int(str(sic)[0]) if str(sic).isdigit() and len(str(sic)) >= 1 else None
    if first is None:
        return None
    return {
        0: "Agriculture",
        1: "Mining",
        2: "Construction",
        3: "Manufacturing",
        4: "Transportation & Utilities",
        5: "Wholesale Trade",
        6: "Retail Trade",
        7: "Services",
        8: "Public Administration",
        9: "Unclassified",
    }.get(first, None)


# -----------------------------
# Config: SIC allowlist (optional)
# -----------------------------

def ensure_sic_config() -> None:
    if not SIC_CONFIG.exists() or SIC_CONFIG.stat().st_size < 5:
        SIC_CONFIG.write_text(
            json.dumps({"sic_allowlist": DEFAULT_SIC_ALLOWLIST}, indent=2),
            encoding="utf-8",
        )


def load_sic_allowlist() -> set[int]:
    ensure_sic_config()
    cfg = json.loads(SIC_CONFIG.read_text(encoding="utf-8"))
    return set(int(x) for x in cfg.get("sic_allowlist", []))


# -----------------------------
# Universe input
# -----------------------------

def is_valid_xlsx(path: Path) -> bool:
    if not path.exists() or path.stat().st_size < 100:
        return False
    try:
        with zipfile.ZipFile(path, "r") as z:
            z.namelist()
        return True
    except Exception:
        return False


def resolve_universe_input() -> Path:
    for p in UNIVERSE_CANDIDATES:
        if p.exists() and p.stat().st_size > 0:
            return p
    raise FileNotFoundError(
        "No universe input found. Create one of: "
        + ", ".join(str(p) for p in UNIVERSE_CANDIDATES[:-1])
        + " (or put tickers into the 'Universe' sheet of the output workbook)."
    )


def dedupe_preserve_order(items: List[str]) -> List[str]:
    seen = set()
    out: List[str] = []
    for x in items:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def load_tickers_from_csv(path: Path) -> List[str]:
    tickers: List[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames and any(fn and fn.strip().lower() == "ticker" for fn in reader.fieldnames):
            for row in reader:
                t = clean_ticker(row.get("ticker") or "")
                if t:
                    tickers.append(t)
        else:
            f.seek(0)
            raw = csv.reader(f)
            for r in raw:
                if not r:
                    continue
                t = clean_ticker(r[0] if r else "")
                if t and t != "TICKER":
                    tickers.append(t)
    return dedupe_preserve_order(tickers)


def load_tickers_from_xlsx(path: Path, sheet_name: str, ticker_header: str) -> List[str]:
    if not is_valid_xlsx(path):
        raise RuntimeError(f"{path} is not a valid .xlsx file.")
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found in {path}")
    ws = wb[sheet_name]

    header_row = None
    header_map: Dict[str, int] = {}

    for r in range(1, min(20, ws.max_row) + 1):
        row_map: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                row_map[normalize_header(v)] = c
        if normalize_header(ticker_header) in row_map:
            header_row = r
            header_map = row_map
            break

    if header_row is None:
        raise RuntimeError(f"Could not find a '{ticker_header}' column in {path} / sheet '{sheet_name}'")

    tcol = header_map[normalize_header(ticker_header)]
    tickers: List[str] = []
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=tcol).value
        if v is None:
            continue
        t = clean_ticker(str(v))
        if t:
            tickers.append(t)

    return dedupe_preserve_order(tickers)


def load_tickers_any(path: Path) -> List[str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_tickers_from_csv(path)
    if suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return load_tickers_from_xlsx(path, UNIVERSE_SHEET, UNIVERSE_TICKER_HEADER)
    raise ValueError(f"Unsupported universe input type: {path}")


# -----------------------------
# SEC fetch helpers (cached, safe)
# -----------------------------

def _sec_http_headers() -> Dict[str, str]:
    return {
        "User-Agent": USER_AGENT,
        "Accept-Encoding": "gzip, deflate",
    }


def get_json_cached(
    url: str,
    cache_key: str,
    *,
    headers: Optional[Dict[str, str]] = None,
    session: Optional[requests.Session] = None,
    sleep_fn = None,
) -> Optional[Dict[str, Any]]:
    cache_path = CACHE_DIR / f"{cache_key}.json"
    if cache_path.exists() and cache_path.stat().st_size > 10:
        try:
            return json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    try:
        sess = session or SEC_SESSION
        hdrs = headers or _sec_http_headers()
        resp = sess.get(url, headers=hdrs, timeout=HTTP_TIMEOUT)
        (sleep_fn or sleep_sec)()
        resp.raise_for_status()
        data = resp.json()
        cache_path.write_text(json.dumps(data), encoding="utf-8")
        return data
    except Exception:
        return None


def build_ticker_to_cik_map() -> Dict[str, Tuple[str, str]]:
    data = get_json_cached(SEC_COMPANY_TICKERS_URL, cache_key="company_tickers")
    mapping: Dict[str, Tuple[str, str]] = {}
    if not data:
        return mapping
    for _, rec in data.items():
        ticker = str(rec.get("ticker", "")).upper().strip()
        cik_str = str(rec.get("cik_str", "")).strip()
        title = str(rec.get("title", "")).strip()
        if ticker and cik_str.isdigit():
            mapping[ticker] = (cik_str.zfill(10), title)
    return mapping


def resolve_cik_from_map(ticker: str, t2c: Dict[str, Tuple[str, str]]) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    for tv in ticker_variants(ticker):
        if tv in t2c:
            cik10, name = t2c[tv]
            return tv, cik10, name
    return None, None, None


def fetch_company_submissions(cik10: str) -> Optional[Dict[str, Any]]:
    url = SEC_SUBMISSIONS_URL_TMPL.format(cik10=cik10)
    return get_json_cached(url, cache_key=f"submissions_{cik10}")


def fetch_company_facts(cik10: str) -> Optional[Dict[str, Any]]:
    url = SEC_COMPANYFACTS_URL_TMPL.format(cik10=cik10)
    return get_json_cached(url, cache_key=f"companyfacts_{cik10}")


def extract_sic(submissions: Dict[str, Any]) -> Tuple[Optional[int], Optional[str]]:
    sic_raw = submissions.get("sic")
    sic_desc = submissions.get("sicDescription")
    sic = None
    if sic_raw is not None:
        try:
            sic = int(str(sic_raw).strip())
        except Exception:
            sic = None
    if sic_desc is not None:
        sic_desc = str(sic_desc).strip()
    return sic, sic_desc


# -----------------------------
# Gate 1: reporting recency (submissions-only)
# -----------------------------

def _latest_date_for_form(submissions: Dict[str, Any], forms_wanted: set[str]) -> Optional[date]:
    recent = (((submissions or {}).get("filings") or {}).get("recent") or {})
    forms = recent.get("form") or []
    filing_dates = recent.get("filingDate") or []

    best: Optional[date] = None
    for i, f in enumerate(forms):
        if f not in forms_wanted:
            continue
        if i >= len(filing_dates):
            continue
        dd = parse_date(str(filing_dates[i]))
        if dd and (best is None or dd > best):
            best = dd
    return best


def gate1_reporting(submissions: Dict[str, Any]) -> Tuple[bool, str, Dict[str, Any]]:
    today = date.today()

    latest_10k = _latest_date_for_form(submissions, {"10-K", "10-K/A"})
    latest_10q = _latest_date_for_form(submissions, {"10-Q", "10-Q/A"})
    latest_20f = _latest_date_for_form(submissions, {"20-F", "20-F/A"})

    pass_10k = bool(latest_10k and (today - latest_10k).days <= G1_MAX_AGE_10K_DAYS)
    pass_10q = bool(latest_10q and (today - latest_10q).days <= G1_MAX_AGE_10Q_DAYS)
    pass_20f = bool(latest_20f and (today - latest_20f).days <= G1_MAX_AGE_20F_DAYS)

    passed = pass_10k or pass_10q or pass_20f

    extras = {
        "G1_Pass_Reporting": "TRUE" if passed else "FALSE",
        "G1_Latest_10K": latest_10k.isoformat() if latest_10k else None,
        "G1_Latest_10Q": latest_10q.isoformat() if latest_10q else None,
        "G1_Latest_20F": latest_20f.isoformat() if latest_20f else None,
    }

    if passed:
        reason = "PASS: Recent filing"
    else:
        if latest_10k or latest_10q or latest_20f:
            reason = "FAIL: Filings exist but stale"
        else:
            reason = "FAIL: No 10-K/10-Q/20-F found"

    extras["G1_Reason"] = reason
    return passed, reason, extras


# -----------------------------
# Companyfacts parsing (concise core metrics only)
# -----------------------------

GAAP_TAGS = {
    "revenue": ["Revenues", "SalesRevenueNet"],
    "op_income": ["OperatingIncomeLoss"],
    "ocf": ["NetCashProvidedByUsedInOperatingActivities"],
    "capex": ["PaymentsToAcquirePropertyPlantAndEquipment"],
    "cash": [
        "CashAndCashEquivalentsAtCarryingValue",
        "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents",
    ],
    "cash_st_inv": ["CashAndCashEquivalentsAndShortTermInvestments"],
    "debt_lt": ["LongTermDebtNoncurrent", "LongTermDebt"],
    "debt_st": ["DebtCurrent", "LongTermDebtCurrent"],
    "assets": ["Assets"],  # Gate 2 uses assets threshold
}

DEI_TAGS = {
    "shares_out": ["EntityCommonStockSharesOutstanding"],
}


def _facts_units(facts: Dict[str, Any], namespace: str, tag: str) -> Optional[Dict[str, Any]]:
    try:
        return facts["facts"][namespace][tag]["units"]
    except Exception:
        return None


def _iter_fact_points(units: Dict[str, Any]) -> Iterable[Dict[str, Any]]:
    for _, arr in units.items():
        if isinstance(arr, list):
            for p in arr:
                if isinstance(p, dict):
                    yield p


def _pick_period_points(
    facts: Dict[str, Any],
    namespace: str,
    tags: List[str],
    *,
    forms: Optional[set[str]] = None,
    fps: Optional[set[str]] = None,
) -> List[Dict[str, Any]]:
    points: List[Dict[str, Any]] = []
    for tag in tags:
        units = _facts_units(facts, namespace, tag)
        if not units:
            continue
        for p in _iter_fact_points(units):
            if forms and p.get("form") not in forms:
                continue
            if fps and p.get("fp") not in fps:
                continue
            if not p.get("end") or p.get("val") is None:
                continue
            points.append(p)
    points.sort(key=lambda x: x.get("end", ""))
    return points


def _latest_instant(points: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    return points[-1] if points else None


def _latest_four_quarters(points: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return points[-4:] if len(points) > 4 else points


def _sum_vals(points: List[Dict[str, Any]]) -> Optional[float]:
    vals = [safe_float(p.get("val")) for p in points]
    vals = [v for v in vals if v is not None]
    if not vals or len(vals) != len(points):
        return None
    return float(sum(vals))


def _yoy_growth(latest: float, prev: float) -> Optional[float]:
    if latest is None or prev is None:
        return None
    if prev == 0:
        return None
    return (latest / prev) - 1.0


def core_metrics_companyfacts(companyfacts: Dict[str, Any]) -> Dict[str, Any]:
    """
    Concise metrics to support:
      - Gate 2 (screenability)
      - Overview (MarketCap comes from Gate 3; everything else is here)

    Produces:
      Overview-facing:
        - TTM_Revenue_USD
        - Revenue_YoY_%
        - Operating_Margin_%
        - Free_Cash_Flow_TTM_USD
        - Cash_USD
        - Total_Debt_USD
        - Shares_Outstanding

      Clutter / trace:
        - A_TTM_OCF_USD
        - A_TTM_Capex_USD
        - A_TTM_FCF_USD
        - A_Total_Assets_USD
        - *_End / *_AsOf dates
    """
    out: Dict[str, Any] = {}

    forms = {"10-Q", "10-K"}
    fps_flow = {"Q1", "Q2", "Q3", "FY"}
    fps_inst = {"Q1", "Q2", "Q3", "FY"}

    rev_points = _pick_period_points(companyfacts, "us-gaap", GAAP_TAGS["revenue"], forms=forms, fps=fps_flow)
    opinc_points = _pick_period_points(companyfacts, "us-gaap", GAAP_TAGS["op_income"], forms=forms, fps=fps_flow)
    ocf_points = _pick_period_points(companyfacts, "us-gaap", GAAP_TAGS["ocf"], forms=forms, fps=fps_flow)
    capex_points = _pick_period_points(companyfacts, "us-gaap", GAAP_TAGS["capex"], forms=forms, fps=fps_flow)

    cash_points = _pick_period_points(companyfacts, "us-gaap", GAAP_TAGS["cash_st_inv"], forms=forms, fps=fps_inst)
    if not cash_points:
        cash_points = _pick_period_points(companyfacts, "us-gaap", GAAP_TAGS["cash"], forms=forms, fps=fps_inst)

    assets_points = _pick_period_points(companyfacts, "us-gaap", GAAP_TAGS["assets"], forms=forms, fps=fps_inst)

    debt_lt_points = _pick_period_points(companyfacts, "us-gaap", GAAP_TAGS["debt_lt"], forms=forms, fps=fps_inst)
    debt_st_points = _pick_period_points(companyfacts, "us-gaap", GAAP_TAGS["debt_st"], forms=forms, fps=fps_inst)

    shares_points = _pick_period_points(companyfacts, "dei", DEI_TAGS["shares_out"], forms=forms, fps=fps_inst)

    # --- TTM Revenue (and end date) ---
    ttm_rev = None
    ttm_rev_end = None
    if len(rev_points) >= 4:
        last4 = _latest_four_quarters(rev_points)
        ttm_rev = _sum_vals(last4)
        ttm_rev_end = last4[-1].get("end")
    else:
        fy = [p for p in rev_points if p.get("fp") == "FY"]
        if fy:
            last = fy[-1]
            ttm_rev = safe_float(last.get("val"))
            ttm_rev_end = last.get("end")

    out["TTM_Revenue_USD"] = ttm_rev
    out["A_TTM_Revenue_End"] = ttm_rev_end

    # --- Revenue YoY (best-effort) ---
    yoy = None
    # Prefer FY YoY if available
    fy_rev = [p for p in rev_points if p.get("fp") == "FY"]
    if len(fy_rev) >= 2:
        yoy = _yoy_growth(safe_float(fy_rev[-1].get("val")), safe_float(fy_rev[-2].get("val")))
    elif len(rev_points) >= 8:
        latest = rev_points[-1]
        prev_year = rev_points[-5]
        yoy = _yoy_growth(safe_float(latest.get("val")), safe_float(prev_year.get("val")))

    out["Revenue_YoY_%"] = yoy

    # --- Operating margin (latest period points) ---
    op_margin = None
    if rev_points and opinc_points:
        rev_last = safe_float(rev_points[-1].get("val"))
        op_last = safe_float(opinc_points[-1].get("val"))
        if rev_last not in (None, 0.0) and op_last is not None:
            op_margin = op_last / rev_last
    out["Operating_Margin_%"] = op_margin

    # --- OCF / Capex / FCF (TTM-ish) ---
    ttm_ocf = None
    if len(ocf_points) >= 4:
        ttm_ocf = _sum_vals(_latest_four_quarters(ocf_points))
    else:
        fy = [p for p in ocf_points if p.get("fp") == "FY"]
        if fy:
            ttm_ocf = safe_float(fy[-1].get("val"))
    out["A_TTM_OCF_USD"] = ttm_ocf

    ttm_capex = None
    if len(capex_points) >= 4:
        ttm_capex = _sum_vals(_latest_four_quarters(capex_points))
    else:
        fy = [p for p in capex_points if p.get("fp") == "FY"]
        if fy:
            ttm_capex = safe_float(fy[-1].get("val"))
    out["A_TTM_Capex_USD"] = ttm_capex

    ttm_fcf = None
    if ttm_ocf is not None and ttm_capex is not None:
        ttm_fcf = ttm_ocf - ttm_capex
    out["A_TTM_FCF_USD"] = ttm_fcf
    out["Free_Cash_Flow_TTM_USD"] = ttm_fcf  # Overview-facing alias (keeps your clutter rule)

    # --- Cash (latest instant) ---
    cash_latest = None
    cash_end = None
    li = _latest_instant(cash_points)
    if li:
        cash_latest = safe_float(li.get("val"))
        cash_end = li.get("end")
    out["Cash_USD"] = cash_latest
    out["A_Cash_End"] = cash_end

    # --- Assets (latest instant; Gate 2 uses it, but it stays clutter) ---
    assets_latest = None
    assets_end = None
    ai = _latest_instant(assets_points)
    if ai:
        assets_latest = safe_float(ai.get("val"))
        assets_end = ai.get("end")
    out["A_Total_Assets_USD"] = assets_latest
    out["A_Total_Assets_End"] = assets_end

    # --- Debt total (latest instant) ---
    dlt = _latest_instant(debt_lt_points)
    dst = _latest_instant(debt_st_points)
    dltv = safe_float(dlt.get("val")) if dlt else None
    dstv = safe_float(dst.get("val")) if dst else None
    debt_total = None
    if dltv is not None or dstv is not None:
        debt_total = (dltv or 0.0) + (dstv or 0.0)
    out["Total_Debt_USD"] = debt_total

    # --- Shares outstanding (latest) ---
    sh = _latest_instant(shares_points)
    out["Shares_Outstanding"] = safe_float(sh.get("val")) if sh else None
    out["A_Shares_AsOf"] = sh.get("end") if sh else None

    return out


# -----------------------------
# Gate 2: minimum screenability (tight, EDGAR-only)
# -----------------------------

def gate2_basics(out: Dict[str, Any]) -> Tuple[bool, str, Dict[str, Any]]:
    """
    Gate 2 (tight, EDGAR-only):
      - (TTM Revenue >= $25m OR Total Assets >= $50m)
      - Shares outstanding > 0
      - Cash >= $5m OR TTM OCF exists
    """
    rev = safe_float(out.get("TTM_Revenue_USD"))
    assets = safe_float(out.get("A_Total_Assets_USD"))
    shares = safe_float(out.get("Shares_Outstanding"))
    cash = safe_float(out.get("Cash_USD"))
    ocf = safe_float(out.get("A_TTM_OCF_USD"))

    size_ok = ((rev is not None and rev >= G2_MIN_TTM_REVENUE_USD) or (assets is not None and assets >= G2_MIN_TOTAL_ASSETS_USD))
    shares_ok = (shares is not None and shares > 0)
    cash_or_ocf_ok = ((cash is not None and cash >= G2_MIN_CASH_USD) or (ocf is not None))

    passed = size_ok and shares_ok and cash_or_ocf_ok

    if not size_ok:
        reason = f"FAIL: Too small (Rev<{G2_MIN_TTM_REVENUE_USD/1e6:.0f}m and Assets<{G2_MIN_TOTAL_ASSETS_USD/1e6:.0f}m) or missing"
    elif not shares_ok:
        reason = "FAIL: Shares outstanding missing/<=0"
    elif not cash_or_ocf_ok:
        reason = f"FAIL: Cash<{G2_MIN_CASH_USD/1e6:.0f}m and OCF missing"
    else:
        reason = "PASS: EDGAR fundamentals screenable (tight)"

    extras = {
        "G2_Pass_Basics": "TRUE" if passed else "FALSE",
        "G2_Reason": reason,
        "G2_Size_OK": "TRUE" if size_ok else "FALSE",
        "G2_Shares_OK": "TRUE" if shares_ok else "FALSE",
        "G2_CashOrOCF_OK": "TRUE" if cash_or_ocf_ok else "FALSE",
    }
    return passed, reason, extras


# -----------------------------
# Gate 3: Market cap + liquidity (Yahoo fetched last, single call)
# -----------------------------

def yahoo_quote_and_liquidity(ticker: str, *, adv_days: int = 30, lookback_days: int = 70) -> Dict[str, Any]:
    """
    One Yahoo chart call (disk-cached per ticker per day):
      - Close_Price_USD (last close)
      - Price_Date
      - ADV30_Shares
      - ADV30_Dollar
    """
    yahoo_ticker = ticker.replace(".", "-").upper()
    end = datetime.now(timezone.utc).date()
    start = end - timedelta(days=lookback_days)

    period1 = int(time.mktime(start.timetuple()))
    period2 = int(time.mktime((end + timedelta(days=1)).timetuple()))

    url = (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{yahoo_ticker}"
        f"?period1={period1}&period2={period2}&interval=1d"
    )

    cache_key = f"yahoo_chart_{yahoo_ticker}_{end.isoformat()}"
    headers = {"User-Agent": "Mozilla/5.0", "Accept": "application/json,text/plain,*/*"}

    j = get_json_cached(
        url,
        cache_key=cache_key,
        headers=headers,
        session=YAHOO_SESSION,
        sleep_fn=sleep_yahoo,
    )

    if not j:
        return {"Close_Price_USD": None, "Price_Date": None, "ADV30_Shares": None, "ADV30_Dollar": None}

    result = (j.get("chart") or {}).get("result") or []
    if not result:
        return {"Close_Price_USD": None, "Price_Date": None, "ADV30_Shares": None, "ADV30_Dollar": None}

    res0 = result[0]
    ts = res0.get("timestamp") or []
    quote0 = (((res0.get("indicators") or {}).get("quote") or [{}])[0])
    closes = quote0.get("close") or []
    vols = quote0.get("volume") or []

    rows: List[Tuple[date, float, float]] = []
    for tstamp, c, v in zip(ts, closes, vols):
        if c is None or v is None:
            continue
        d = datetime.fromtimestamp(int(tstamp), tz=timezone.utc).date()
        rows.append((d, float(c), float(v)))

    if not rows:
        return {"Close_Price_USD": None, "Price_Date": None, "ADV30_Shares": None, "ADV30_Dollar": None}

    rows.sort(key=lambda x: x[0])
    last_d, last_c, _ = rows[-1]

    tail = rows[-adv_days:] if len(rows) >= adv_days else rows
    adv_sh = (sum(r[2] for r in tail) / len(tail)) if tail else None
    adv_dl = (sum((r[1] * r[2]) for r in tail) / len(tail)) if tail else None

    return {
        "Close_Price_USD": last_c,
        "Price_Date": last_d.isoformat(),
        "ADV30_Shares": adv_sh,
        "ADV30_Dollar": adv_dl,
    }


def gate3_marketcap_and_liquidity(out: Dict[str, Any], ticker: str) -> Tuple[bool, str, Dict[str, Any]]:
    shares = safe_float(out.get("Shares_Outstanding"))
    q = yahoo_quote_and_liquidity(ticker)

    price = safe_float(q.get("Close_Price_USD"))
    adv_sh = safe_float(q.get("ADV30_Shares"))
    adv_dl = safe_float(q.get("ADV30_Dollar"))

    mcap = None
    if shares is not None and price is not None:
        mcap = shares * price

    mcap_ok = (mcap is not None and mcap <= G3_MAX_MARKET_CAP_USD)
    liq_ok = (
        adv_sh is not None and adv_sh >= G3_MIN_ADV30_SHARES
        and adv_dl is not None and adv_dl >= G3_MIN_ADV30_DOLLAR
    )

    passed = mcap_ok and liq_ok

    if mcap is None:
        reason = "FAIL: Could not compute market cap (missing price or shares)"
    elif not mcap_ok:
        reason = f"FAIL: Market cap > ${G3_MAX_MARKET_CAP_USD:,.0f}"
    elif not liq_ok:
        reason = f"FAIL: Illiquid (ADV30<{G3_MIN_ADV30_SHARES:,} or $ADV30<${G3_MIN_ADV30_DOLLAR:,.0f})"
    else:
        reason = "PASS: Market cap + liquidity (tight)"

    # Overview-facing: Market cap + ADV30 shares
    extras = {
        "G3_Pass": "TRUE" if passed else "FALSE",
        "G3_Reason": reason,

        "Market_Cap_USD": mcap,
        "ADV30_Shares": adv_sh,

        # Clutter mechanics (per your rule)
        "P_Close_Price_USD": q.get("Close_Price_USD"),
        "P_Price_Date": q.get("Price_Date"),
        "P_ADV30_Dollar": adv_dl,
        "P_MarketCap_Method": "Yahoo Last Close × EDGAR Shares" if mcap is not None else None,
    }
    return passed, reason, extras


# -----------------------------
# Excel writing (robust, update-by-ticker)
# -----------------------------
# Visible structure: Intro + Overview, then Clutter.

INTRO_HEADERS = [
    "Ticker",
    "Company Name",
    "Outcome",
    "Sector",
    "Industry",
]

OVERVIEW_HEADERS = [
    "Market_Cap_USD",
    "TTM_Revenue_USD",
    "Revenue_YoY_%",
    "Operating_Margin_%",
    "Free_Cash_Flow_TTM_USD",
    "Cash_USD",
    "Total_Debt_USD",
    "Shares_Outstanding",
    "ADV30_Shares",
]

CLUTTER_HEADERS = [
    # Identity + classification mechanics
    "CIK",
    "SIC",
    "SIC Description",

    # Run metadata
    "Data As-Of Date",
    "Run Timestamp (UTC)",
    "Status",

    # Gate 1
    "G1_Pass_Reporting",
    "G1_Reason",
    "G1_Latest_10K",
    "G1_Latest_10Q",
    "G1_Latest_20F",

    # Gate 2
    "G2_Pass_Basics",
    "G2_Reason",
    "G2_Size_OK",
    "G2_Shares_OK",
    "G2_CashOrOCF_OK",

    # Gate 3
    "G3_Pass",
    "G3_Reason",

    # Price/quote mechanics (clutter per your rule)
    "P_Close_Price_USD",
    "P_Price_Date",
    "P_ADV30_Dollar",
    "P_MarketCap_Method",

    # Internals (clutter per your rule)
    "A_TTM_Revenue_End",
    "A_Total_Assets_USD",
    "A_Total_Assets_End",
    "A_TTM_OCF_USD",
    "A_TTM_Capex_USD",
    "A_TTM_FCF_USD",
    "A_Cash_End",
    "A_Shares_AsOf",
]

BASE_HEADERS = INTRO_HEADERS + OVERVIEW_HEADERS + CLUTTER_HEADERS


def ensure_output_workbook() -> None:
    if is_valid_xlsx(OUTPUT_XLSX):
        wb = load_workbook(OUTPUT_XLSX)
        if UNIVERSE_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(UNIVERSE_SHEET)
            ws.append(BASE_HEADERS)
            wb.save(OUTPUT_XLSX)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = UNIVERSE_SHEET
    ws.append(BASE_HEADERS)
    wb.save(OUTPUT_XLSX)


def find_header_row(ws: Worksheet, max_rows: int = 20) -> Tuple[int, Dict[str, int]]:
    best_row = None
    best_map: Dict[str, int] = {}

    for r in range(1, max_rows + 1):
        row_map: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                row_map[normalize_header(v)] = c
        if "ticker" in row_map:
            best_row = r
            best_map = row_map
            break

    if best_row is None:
        best_row = 1
        best_map = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if isinstance(v, str) and v.strip():
                best_map[normalize_header(v)] = c

    return best_row, best_map


def ensure_columns(ws: Worksheet, header_row: int, header_map: Dict[str, int], columns: List[str]) -> Dict[str, int]:
    max_col = ws.max_column
    for col_name in columns:
        key = normalize_header(col_name)
        if key in header_map:
            continue
        max_col += 1
        ws.cell(row=header_row, column=max_col).value = col_name
        header_map[key] = max_col
    return header_map


def build_ticker_row_index(ws: Worksheet, header_row: int, ticker_col: int) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    r = header_row + 1
    blanks_in_a_row = 0
    while r <= ws.max_row:
        v = ws.cell(row=r, column=ticker_col).value
        if v in (None, ""):
            blanks_in_a_row += 1
            if blanks_in_a_row >= 50:   # stop after a reasonable blank run
                break
        else:
            blanks_in_a_row = 0
            t = str(v).strip().upper()
            if t:
                idx[t] = r
        r += 1
    return idx


def write_company_dicts_to_excel(xlsx_path: Path, rows: List[Dict[str, Any]]) -> None:
    wb = load_workbook(xlsx_path)
    if UNIVERSE_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(UNIVERSE_SHEET)
        ws.append(BASE_HEADERS)
    ws = wb[UNIVERSE_SHEET]

    header_row, header_map = find_header_row(ws)
    header_map = ensure_columns(ws, header_row, header_map, BASE_HEADERS)

    ticker_col = header_map.get("ticker")
    if not ticker_col:
        header_map = ensure_columns(ws, header_row, header_map, ["Ticker"])
        ticker_col = header_map["ticker"]

    existing = build_ticker_row_index(ws, header_row, ticker_col)

    append_row = header_row + 1
    while True:
        v = ws.cell(row=append_row, column=ticker_col).value
        if v in (None, ""):
            break
        append_row += 1


    for rd in rows:
        ticker = str(rd.get("Ticker") or "").strip().upper()
        if not ticker:
            continue

        row_num = existing.get(ticker)
        if row_num is None:
            row_num = append_row
            append_row += 1
            existing[ticker] = row_num

        # Only write known headers by default (prevents random accidental keys exploding columns)
        for k in BASE_HEADERS:
            v = rd.get(k)
            col = header_map.get(normalize_header(k))
            if col:
                ws.cell(row=row_num, column=col).value = na(v)

    wb.save(xlsx_path)


# -----------------------------
# Main orchestration
# -----------------------------

def build_base_row(
    ticker: str,
    *,
    company_name: Optional[str],
    cik10: Optional[str],
    sic: Optional[int],
    sic_desc: Optional[str],
    status: str,
    as_of: str,
) -> Dict[str, Any]:
    sector = sic_to_sector(sic)
    industry = sic_desc

    # NEW: short, human-readable outcome (Intro)
    outcome = status
    if isinstance(status, str):
        if status.startswith("Excluded (Gate 1)"):
            outcome = "Excluded (G1)"
        elif status.startswith("Excluded (Gate 2)"):
            outcome = "Excluded (G2)"
        elif status.startswith("Excluded (Gate 3)"):
            outcome = "Excluded (G3)"
        elif status == "OK":
            outcome = "PASS"
        elif "Ticker not found" in status:
            outcome = "No CIK"

    return {
        # Intro
        "Ticker": ticker,
        "Company Name": company_name,
        "Outcome": outcome,   # NEW
        "Sector": sector,
        "Industry": industry,

        # Clutter identity fields
        "CIK": cik10,
        "SIC": sic,
        "SIC Description": sic_desc,

        # Metadata
        "Data As-Of Date": as_of,
        "Run Timestamp (UTC)": now_iso(),
        "Status": status,
    }

def outcome_from_status(status: Optional[str]) -> Optional[str]:
    s = (status or "").strip()
    if not s:
        return None
    if s.startswith("Excluded (Gate 1)"):
        return "Excluded (G1)"
    if s.startswith("Excluded (Gate 2)"):
        return "Excluded (G2)"
    if s.startswith("Excluded (Gate 3)"):
        return "Excluded (G3)"
    if "Ticker not found" in s:
        return "No CIK"
    if s == "No submissions JSON":
        return "No submissions"
    if s == "OK":
        return "PASS"
    if "Failed" in s or "error" in s.lower():
        return "Error"
    return s


def main() -> None:
    ensure_output_workbook()

    universe_path = resolve_universe_input()
    tickers = load_tickers_any(universe_path)
    if not tickers:
        raise RuntimeError(f"No tickers found in universe input: {universe_path}")

    t2c = build_ticker_to_cik_map()
    as_of = date.today().isoformat()

    out_rows: List[Dict[str, Any]] = []

    for t in tickers:
        t = (t or "").strip().upper()
        if not t:
            continue

        _, cik10, name = resolve_cik_from_map(t, t2c)

        # --- Ticker not mapped: immediate stop ---
        if not cik10:
            out = build_base_row(
                ticker=t,
                company_name=None,
                cik10=None,
                sic=None,
                sic_desc=None,
                status="Ticker not found in SEC ticker map",
                as_of=as_of,
            )
            # gates (clutter)
            out.update({
                "G1_Pass_Reporting": "FALSE",
                "G1_Reason": "FAIL: Not in SEC ticker map",
                "G2_Pass_Basics": "FALSE",
                "G2_Reason": "FAIL: Gate 1 failed (no CIK)",
                "G3_Pass": "FALSE",
                "G3_Reason": "FAIL: Gate 1 failed (no CIK)",
            })
            out_rows.append(out)
            continue

        try:
            submissions = fetch_company_submissions(cik10)
            if not submissions:
                out = build_base_row(
                    ticker=t,
                    company_name=name,
                    cik10=cik10,
                    sic=None,
                    sic_desc=None,
                    status="No submissions JSON",
                    as_of=as_of,
                )
                out.update({
                    "G1_Pass_Reporting": "FALSE",
                    "G1_Reason": "FAIL: No submissions JSON",
                    "G2_Pass_Basics": "FALSE",
                    "G2_Reason": "FAIL: Gate 1 failed (no submissions)",
                    "G3_Pass": "FALSE",
                    "G3_Reason": "FAIL: Gate 1 failed (no submissions)",
                })
                out_rows.append(out)
                continue

            sic, sic_desc = extract_sic(submissions)

            out = build_base_row(
                ticker=t,
                company_name=name,
                cik10=cik10,
                sic=sic,
                sic_desc=sic_desc,
                status="OK",
                as_of=as_of,
            )

            # -----------------------------
            # Gate 1
            # -----------------------------
            if ENABLE_GATE_1:
                g1_passed, _, g1_fields = gate1_reporting(submissions)
                out.update(g1_fields)
                if not g1_passed:
                    out["Status"] = "Excluded (Gate 1)"
                    out.update({
                        "G2_Pass_Basics": "FALSE",
                        "G2_Reason": "FAIL: Gate 1 failed",
                        "G3_Pass": "FALSE",
                        "G3_Reason": "FAIL: Gate 1 failed",
                    })
                    out_rows.append(out)
                    continue

            # -----------------------------
            # Gate 2 (EDGAR-only, tight)
            # -----------------------------
            facts = fetch_company_facts(cik10)
            if not facts:
                out["Status"] = "Excluded (Gate 2)"
                out.update({
                    "G2_Pass_Basics": "FALSE",
                    "G2_Reason": "FAIL: Missing companyfacts",
                    "G3_Pass": "FALSE",
                    "G3_Reason": "FAIL: Gate 2 failed",
                })
                out_rows.append(out)
                continue

            # Parse only concise core metrics
            try:
                core = core_metrics_companyfacts(facts)
                out.update(core)

                # Populate overview keys from core (no duplicates)
                out["TTM_Revenue_USD"] = core.get("TTM_Revenue_USD")
                out["Revenue_YoY_%"] = core.get("Revenue_YoY_%")
                out["Operating_Margin_%"] = core.get("Operating_Margin_%")
                out["Free_Cash_Flow_TTM_USD"] = core.get("Free_Cash_Flow_TTM_USD")
                out["Cash_USD"] = core.get("Cash_USD")
                out["Total_Debt_USD"] = core.get("Total_Debt_USD")
                out["Shares_Outstanding"] = core.get("Shares_Outstanding")
            except Exception:
                out["Status"] = "Excluded (Gate 2)"
                out.update({
                    "G2_Pass_Basics": "FALSE",
                    "G2_Reason": "FAIL: Could not parse core fundamentals",
                    "G3_Pass": "FALSE",
                    "G3_Reason": "FAIL: Gate 2 failed",
                })
                out_rows.append(out)
                continue

            g2_passed, _, g2_fields = gate2_basics(out)
            out.update(g2_fields)

            if ENABLE_GATE_2 and not g2_passed:
                out["Status"] = "Excluded (Gate 2)"
                out.update({
                    "G3_Pass": "FALSE",
                    "G3_Reason": "FAIL: Gate 2 failed",
                })
                out_rows.append(out)
                continue

            # -----------------------------
            # Gate 3 (Yahoo fetched LAST)
            # -----------------------------
            if ENABLE_GATE_3:
                g3_passed, _, g3_fields = gate3_marketcap_and_liquidity(out, t)
                out.update(g3_fields)
                if not g3_passed:
                    out["Status"] = "Excluded (Gate 3)"
                    out_rows.append(out)
                    continue

            out["Status"] = "OK"

            out_rows.append(out)

        except Exception:
            out = build_base_row(
                ticker=t,
                company_name=name,
                cik10=cik10,
                sic=None,
                sic_desc=None,
                status="Failed (unexpected error in company loop)",
                as_of=as_of,
            )
            out.update({
                "G1_Pass_Reporting": "FALSE",
                "G1_Reason": "FAIL: Unexpected error",
                "G2_Pass_Basics": "FALSE",
                "G2_Reason": "FAIL: Unexpected error",
                "G3_Pass": "FALSE",
                "G3_Reason": "FAIL: Unexpected error",
            })
            out_rows.append(out)

    write_company_dicts_to_excel(OUTPUT_XLSX, out_rows)

    print(f"Universe input: {universe_path}")
    print(f"Processed tickers: {len(tickers)}")
    print(f"Updated workbook: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
