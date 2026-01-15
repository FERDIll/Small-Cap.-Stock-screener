#!/usr/bin/env python3
"""
v1 SEC scraper: tickers -> CIK -> SIC -> write to Excel template.

- Inputs:
  - data/universe.csv (tickers)
  - config/sic_aero_defense.json (SIC allowlist)
  - data/defense_screening_prototype_v1.xlsx (template)

- Output:
  - Overwrites template in-place (or writes to a new file if you set OUTPUT_XLSX)

Notes:
- Be a good citizen: set a real User-Agent with contact info and rate-limit requests.
"""

from __future__ import annotations

import csv
import json
import time
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# -----------------------------
# Configuration
# -----------------------------

REPO_ROOT = Path(__file__).resolve().parents[1]

UNIVERSE_CSV = REPO_ROOT / "data" / "universe.csv"
TEMPLATE_XLSX = REPO_ROOT / "data" / "defense_screening_prototype_v1.xlsx"
OUTPUT_XLSX = TEMPLATE_XLSX  # set to REPO_ROOT/"data"/"output.xlsx" if you want a separate file
SIC_CONFIG = REPO_ROOT / "config" / "sic_aero_defense.json"

# SEC endpoints
SEC_COMPANY_TICKERS_URL = "https://www.sec.gov/files/company_tickers.json"
SEC_SUBMISSIONS_URL_TMPL = "https://data.sec.gov/submissions/CIK{cik10}.json"

# IMPORTANT: Put real contact info here (SEC expects identifiable User-Agent).
USER_AGENT = "YourName YourProject (your.email@example.com)"

# Rate limiting: keep it conservative for a manual monthly run
SECONDS_BETWEEN_REQUESTS = 0.2

# Cache folder (optional but helpful)
CACHE_DIR = REPO_ROOT / ".cache" / "sec"
CACHE_DIR.mkdir(parents=True, exist_ok=True)


# -----------------------------
# Helpers
# -----------------------------

def _sleep():
    time.sleep(SECONDS_BETWEEN_REQUESTS)

def _get_json(url: str, cache_key: str) -> Dict[str, Any]:
    """
    Fetch JSON with simple file cache.
    """
    cache_path = CACHE_DIR / f"{cache_key}.json"
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))

    headers = {
        "User-Agent": USER_AGENT,
        "Accept-Encoding": "gzip, deflate",
        "Host": "www.sec.gov" if "sec.gov/files" in url else "data.sec.gov",
    }

    resp = requests.get(url, headers=headers, timeout=30)
    _sleep()
    resp.raise_for_status()

    data = resp.json()
    cache_path.write_text(json.dumps(data), encoding="utf-8")
    return data

def load_sic_allowlist(path: Path) -> set[int]:
    cfg = json.loads(path.read_text(encoding="utf-8"))
    return set(int(x) for x in cfg.get("sic_allowlist", []))

def load_tickers(path: Path) -> List[str]:
    """
    Accepts:
    - one-column CSV with header 'ticker'
    - or a CSV with first column as tickers
    """
    tickers: List[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames and any(fn.lower() == "ticker" for fn in reader.fieldnames):
            for row in reader:
                t = (row.get("ticker") or "").strip().upper()
                if t:
                    tickers.append(t)
        else:
            # fallback: treat as simple CSV without header
            f.seek(0)
            raw = csv.reader(f)
            for r in raw:
                if not r:
                    continue
                t = (r[0] or "").strip().upper()
                if t and t != "TICKER":
                    tickers.append(t)

    # de-dupe while preserving order
    seen = set()
    out = []
    for t in tickers:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out

@dataclass
class CompanyRow:
    ticker: str
    cik: str
    name: str
    sic: Optional[int]
    sic_desc: Optional[str]
    is_aero_defense: bool
    data_as_of: str  # ISO date string


def build_ticker_to_cik_map() -> Dict[str, Tuple[str, str]]:
    """
    Returns {TICKER: (CIK10, COMPANY_NAME)} from SEC's company_tickers.json.
    """
    data = _get_json(SEC_COMPANY_TICKERS_URL, cache_key="company_tickers")
    mapping: Dict[str, Tuple[str, str]] = {}

    # Structure is typically: {"0":{"cik_str":..., "ticker":..., "title":...}, ...}
    for _, rec in data.items():
        ticker = str(rec.get("ticker", "")).upper().strip()
        if not ticker:
            continue
        cik_str = str(rec.get("cik_str", "")).strip()
        title = str(rec.get("title", "")).strip()
        if cik_str.isdigit():
            cik10 = cik_str.zfill(10)
            mapping[ticker] = (cik10, title)

    return mapping

def fetch_company_submissions(cik10: str) -> Dict[str, Any]:
    url = SEC_SUBMISSIONS_URL_TMPL.format(cik10=cik10)
    return _get_json(url, cache_key=f"submissions_{cik10}")

def extract_sic(submissions: Dict[str, Any]) -> Tuple[Optional[int], Optional[str]]:
    sic_raw = submissions.get("sic")
    sic_desc = submissions.get("sicDescription")
    sic = None
    if sic_raw is not None:
        try:
            sic = int(str(sic_raw).strip())
        except ValueError:
            sic = None
    if sic_desc is not None:
        sic_desc = str(sic_desc).strip()
    return sic, sic_desc


# -----------------------------
# Excel writing
# -----------------------------

def find_header_row(ws: Worksheet, max_rows: int = 20) -> Tuple[int, Dict[str, int]]:
    """
    Finds the header row by scanning the first N rows and returning a map
    of normalized header -> column index.
    """
    def norm(s: str) -> str:
        return " ".join(s.strip().lower().split())

    best_row = None
    best_map: Dict[str, int] = {}

    for r in range(1, max_rows + 1):
        row_map: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                row_map[norm(v)] = c
        # Heuristic: header row has at least 3 known fields
        known = sum(1 for k in row_map.keys() if k in {
            "ticker", "symbol", "cik", "company name", "name", "sic", "sic description"
        })
        if known >= 3 and len(row_map) > len(best_map):
            best_row = r
            best_map = row_map

    if best_row is None:
        # Assume first row is header if nothing found
        best_row = 1
        best_map = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if isinstance(v, str) and v.strip():
                best_map[" ".join(v.strip().lower().split())] = c

    return best_row, best_map

def col_for(header_map: Dict[str, int], *candidates: str) -> Optional[int]:
    """
    Find a column index by trying multiple header name candidates.
    """
    def norm(s: str) -> str:
        return " ".join(s.strip().lower().split())
    for cand in candidates:
        c = header_map.get(norm(cand))
        if c:
            return c
    return None

def write_rows_to_excel(xlsx_in: Path, xlsx_out: Path, rows: List[CompanyRow]) -> None:
    wb = load_workbook(xlsx_in)
    if "Universe" not in wb.sheetnames:
        raise RuntimeError("Template workbook must have a sheet named 'Universe'.")

    ws = wb["Universe"]
    header_row, header_map = find_header_row(ws)

    # Resolve columns (tolerant to minor header variations)
    c_ticker = col_for(header_map, "Ticker", "Symbol")
    c_name = col_for(header_map, "Company Name", "Name", "Company")
    c_cik = col_for(header_map, "CIK")
    c_sic = col_for(header_map, "SIC")
    c_sic_desc = col_for(header_map, "SIC Description", "SIC Desc", "SIC Industry")
    c_flag = col_for(header_map, "Aerospace & Defense", "Aerospace and Defense", "A&D", "Aerospace & Defense?")
    c_asof = col_for(header_map, "Data As-Of Date", "Data As Of", "As Of")

    # Determine where to start writing: first empty row after header
    start_row = header_row + 1
    while ws.cell(row=start_row, column=1).value not in (None, ""):
        start_row += 1

    for i, r in enumerate(rows):
        rr = start_row + i
        if c_ticker: ws.cell(row=rr, column=c_ticker).value = r.ticker
        if c_name: ws.cell(row=rr, column=c_name).value = r.name
        if c_cik: ws.cell(row=rr, column=c_cik).value = r.cik
        if c_sic: ws.cell(row=rr, column=c_sic).value = r.sic
        if c_sic_desc: ws.cell(row=rr, column=c_sic_desc).value = r.sic_desc
        if c_flag: ws.cell(row=rr, column=c_flag).value = "TRUE" if r.is_aero_defense else "FALSE"
        if c_asof: ws.cell(row=rr, column=c_asof).value = r.data_as_of

    wb.save(xlsx_out)


# -----------------------------
# Main
# -----------------------------

def main() -> None:
    if not UNIVERSE_CSV.exists():
        raise FileNotFoundError(f"Missing {UNIVERSE_CSV}. Create it with a 'ticker' column.")
    if not TEMPLATE_XLSX.exists():
        raise FileNotFoundError(f"Missing {TEMPLATE_XLSX}. Put your Excel template there.")
    if not SIC_CONFIG.exists():
        raise FileNotFoundError(f"Missing {SIC_CONFIG}. Put sic_allowlist JSON there.")

    sic_allow = load_sic_allowlist(SIC_CONFIG)
    tickers = load_tickers(UNIVERSE_CSV)
    t2c = build_ticker_to_cik_map()

    out_rows: List[CompanyRow] = []
    as_of = date.today().isoformat()

    for t in tickers:
        if t not in t2c:
            # skip unknown tickers (could be non-SEC filer)
            continue

        cik10, name = t2c[t]
        subs = fetch_company_submissions(cik10)
        sic, sic_desc = extract_sic(subs)
        is_ad = (sic in sic_allow) if sic is not None else False

        out_rows.append(
            CompanyRow(
                ticker=t,
                cik=cik10,
                name=name,
                sic=sic,
                sic_desc=sic_desc,
                is_aero_defense=is_ad,
                data_as_of=as_of,
            )
        )

    # Write into the Excel template
    write_rows_to_excel(TEMPLATE_XLSX, OUTPUT_XLSX, out_rows)
    print(f"Wrote {len(out_rows)} rows to {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
