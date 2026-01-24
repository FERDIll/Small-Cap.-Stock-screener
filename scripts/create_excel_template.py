from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule

REPO_ROOT = Path(__file__).resolve().parents[1]
out = REPO_ROOT / "data" / "defense_screening_prototype_v1.xlsx"
out.parent.mkdir(parents=True, exist_ok=True)

# --- Headers (new data flow) ---
INTRO_HEADERS = [
    "Ticker",
    "Company Name",
    "Outcome",
    "Sector",
    "Industry",
]

OVERVIEW_HEADERS = [
    "Market_Cap_USD",
    "TTM_Revenue_USD",
    "Revenue_YoY_%",
    "Operating_Margin_%",
    "Free_Cash_Flow_TTM_USD",
    "Cash_USD",
    "Total_Debt_USD",
    "Shares_Outstanding",
    "ADV30_Shares",
]

CLUTTER_HEADERS = [
    # Identity / classification mechanics
    "CIK",
    "SIC",
    "SIC Description",

    # Metadata
    "Status",
    "Data As-Of Date",
    "Run Timestamp (UTC)",

    # Gate 1
    "G1_Pass_Reporting",
    "G1_Reason",
    "G1_Latest_10K",
    "G1_Latest_10Q",
    "G1_Latest_20F",

    # Gate 2
    "G2_Pass_Basics",
    "G2_Reason",
    "G2_Size_OK",
    "G2_Shares_OK",
    "G2_CashOrOCF_OK",

    # Gate 3
    "G3_Pass",
    "G3_Reason",

    # Price / quote mechanics (clutter per spec)
    "P_Close_Price_USD",
    "P_Price_Date",
    "P_ADV30_Dollar",
    "P_MarketCap_Method",

    # Internals / trace
    "A_TTM_Revenue_End",
    "A_Total_Assets_USD",
    "A_Total_Assets_End",
    "A_TTM_OCF_USD",
    "A_TTM_Capex_USD",
    "A_TTM_FCF_USD",
    "A_Cash_End",
    "A_Shares_AsOf",
]

# Optional (kept far right; hidden)
OPTIONAL_FAR_RIGHT = [
    "C_10Q_Count_365d","C_10K_Count_365d","C_8K_Count_365d","C_S1_or_S3_Count_365d","C_424B_Count_365d",
    "C_13D_13G_Count_365d","C_Form4_Count_90d","C_Latest_10Q_Date","C_Latest_10K_Date","C_Latest_Shelf_Date",
    "C_Latest_Form4_Date","C_Recent_Filings_Listed",
    "B_Form4_Net_Shares_180d","B_Form4_Tx_Count_180d","B_Form4_Last_Date_Parsed",
    "F_TTM_Revenue_lt_500m","F_YoY_Abs_gt_40pct","F_OpMargin_lt_neg10pct","F_RnD_gt_20pct","F_SGandA_gt_50pct",
    "F_CurrentRatio_lt_1_5","F_DebtEq_gt_1_0",
]

HEADERS = INTRO_HEADERS + OVERVIEW_HEADERS + CLUTTER_HEADERS + OPTIONAL_FAR_RIGHT
COL = {h: i + 1 for i, h in enumerate(HEADERS)}

wb = Workbook()
ws = wb.active
ws.title = "Universe"
ws.append(HEADERS)

# --- Header style (readability-first) ---
ws.row_dimensions[1].height = 30
header_fill = PatternFill("solid", fgColor="111827")
header_font = Font(bold=True, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for c in range(1, len(HEADERS) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

# Freeze panes: keep Intro visible while scrolling
ws.freeze_panes = "F2"  # freezes A-D + row 1

# Filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

# --- Column widths ---
def set_w(name, w):
    if name in COL:
        ws.column_dimensions[get_column_letter(COL[name])].width = w

# Intro
set_w("Ticker", 10)
set_w("Company Name", 34)
set_w("Outcome", 14)
set_w("Sector", 20)
set_w("Industry", 26)

# Overview
for n in ["Market_Cap_USD","TTM_Revenue_USD","Free_Cash_Flow_TTM_USD","Cash_USD","Total_Debt_USD","Shares_Outstanding"]:
    set_w(n, 18)
set_w("ADV30_Shares", 14)
set_w("Revenue_YoY_%", 14)
set_w("Operating_Margin_%", 16)

# Clutter & reasons
set_w("Status", 18)
set_w("SIC Description", 26)
for n in ["G1_Reason","G2_Reason","G3_Reason"]:
    set_w(n, 40)
for n in ["P_MarketCap_Method","Run Timestamp (UTC)"]:
    set_w(n, 26)

# Optional far right
for n in OPTIONAL_FAR_RIGHT:
    set_w(n, 16)

# --- Number formats ---
CURRENCY_FMT = '$#,##0'
CURRENCY_FMT_2 = '$#,##0.00'
PCT_FMT = '0.0%'
INT_FMT = '#,##0'
DATE_FMT = 'yyyy-mm-dd'
DATETIME_FMT = 'yyyy-mm-dd hh:mm'

currency_cols_0 = {"Market_Cap_USD","TTM_Revenue_USD","Free_Cash_Flow_TTM_USD","Cash_USD","Total_Debt_USD",
                   "P_ADV30_Dollar","A_Total_Assets_USD","A_TTM_OCF_USD","A_TTM_Capex_USD","A_TTM_FCF_USD"}
currency_cols_2 = {"P_Close_Price_USD"}
pct_cols = {"Revenue_YoY_%","Operating_Margin_%"}
int_cols = {"Shares_Outstanding","ADV30_Shares","C_10Q_Count_365d","C_10K_Count_365d","C_8K_Count_365d",
            "C_S1_or_S3_Count_365d","C_424B_Count_365d","C_13D_13G_Count_365d","C_Form4_Count_90d",
            "C_Recent_Filings_Listed","B_Form4_Net_Shares_180d","B_Form4_Tx_Count_180d"}
date_cols = {"P_Price_Date","G1_Latest_10K","G1_Latest_10Q","G1_Latest_20F","A_TTM_Revenue_End",
             "A_Total_Assets_End","A_Cash_End","A_Shares_AsOf",
             "C_Latest_10Q_Date","C_Latest_10K_Date","C_Latest_Shelf_Date","C_Latest_Form4_Date","B_Form4_Last_Date_Parsed",
             "Data As-Of Date"}
datetime_cols = {"Run Timestamp (UTC)"}

max_row = 5000
for name in HEADERS:
    col = COL[name]
    fmt = None
    if name in currency_cols_0:
        fmt = CURRENCY_FMT
    elif name in currency_cols_2:
        fmt = CURRENCY_FMT_2
    elif name in pct_cols:
        fmt = PCT_FMT
    elif name in int_cols:
        fmt = INT_FMT
    elif name in date_cols:
        fmt = DATE_FMT
    elif name in datetime_cols:
        fmt = DATETIME_FMT

    if fmt:
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=col).number_format = fmt

# --- Grouping / collapsing ---
clutter_start = COL[CLUTTER_HEADERS[0]]
clutter_end = COL[CLUTTER_HEADERS[-1]]
ws.column_dimensions.group(get_column_letter(clutter_start), get_column_letter(clutter_end), hidden=True)

opt_start = COL[OPTIONAL_FAR_RIGHT[0]]
opt_end = COL[OPTIONAL_FAR_RIGHT[-1]]
ws.column_dimensions.group(get_column_letter(opt_start), get_column_letter(opt_end), hidden=True)

ws.sheet_properties.outlinePr.summaryBelow = False

# --- Conditional formatting ---
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
ORANGE = PatternFill("solid", fgColor="FFEB9C")

def col_rng(name: str) -> str:
    c = get_column_letter(COL[name])
    return f"{c}2:{c}{max_row}"

# Status: OK green, Excluded red
c = get_column_letter(COL["Status"])
ws.conditional_formatting.add(col_rng("Status"), FormulaRule(formula=[f'ISNUMBER(SEARCH("OK",${c}2))'], fill=GREEN))
ws.conditional_formatting.add(col_rng("Status"), FormulaRule(formula=[f'ISNUMBER(SEARCH("Excluded",${c}2))'], fill=RED))

# Gate 3 pass
c = get_column_letter(COL["G3_Pass"])
ws.conditional_formatting.add(col_rng("G3_Pass"), FormulaRule(formula=[f'${c}2="TRUE"'], fill=GREEN))
ws.conditional_formatting.add(col_rng("G3_Pass"), FormulaRule(formula=[f'${c}2="FALSE"'], fill=RED))

# Market cap: <1B green, >5B orange
ws.conditional_formatting.add(col_rng("Market_Cap_USD"), CellIsRule(operator="lessThan", formula=["1000000000"], fill=GREEN))
ws.conditional_formatting.add(col_rng("Market_Cap_USD"), CellIsRule(operator="greaterThan", formula=["5000000000"], fill=ORANGE))

# YoY: >30% green, <0 red
ws.conditional_formatting.add(col_rng("Revenue_YoY_%"), CellIsRule(operator="greaterThan", formula=["0.30"], fill=GREEN))
ws.conditional_formatting.add(col_rng("Revenue_YoY_%"), CellIsRule(operator="lessThan", formula=["0"], fill=RED))

# Operating margin: >10% green, <0 red
ws.conditional_formatting.add(col_rng("Operating_Margin_%"), CellIsRule(operator="greaterThan", formula=["0.10"], fill=GREEN))
ws.conditional_formatting.add(col_rng("Operating_Margin_%"), CellIsRule(operator="lessThan", formula=["0"], fill=RED))

# ADV30 shares: <200k red, >1m green
ws.conditional_formatting.add(col_rng("ADV30_Shares"), CellIsRule(operator="lessThan", formula=["200000"], fill=RED))
ws.conditional_formatting.add(col_rng("ADV30_Shares"), CellIsRule(operator="greaterThan", formula=["1000000"], fill=GREEN))

wb.save(out)
print("Created template:", out.resolve())
